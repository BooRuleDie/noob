import random
import os
import platform
import time
from openpyxl import *
from openpyxl.styles import *
from datetime import date
from pathlib import Path
'''
COLOR_CODES: 

Black: 30
Red: 31
Green: 32
Yellow: 33
Blue: 34
Purple: 35
Cyan: 36
White: 37

BACKGRUND_CODES:

COLOR_CODES + 10

Example: \033[2;31m+\033[0;0m prints red plus sign '+'
'''


# FUNCTION DEFINITIONS

def clear_screen(time_to_wait):

    time.sleep(time_to_wait)
    
    Current_OS = platform.system()

    # It'also cross platform :D

    if "windows" in Current_OS.lower():
        os.system("cls")
    else: 
        os.system("clear")

def print_banner():

    banner = """


\033[2;31m██╗     ██╗███╗   ██╗██╗  ██╗██╗███╗   ██╗ ██████╗\033[0;0m     ███╗   ███╗███████╗███╗   ███╗ ██████╗ ██████╗ ██╗   ██╗
\033[2;31m██║     ██║████╗  ██║██║ ██╔╝██║████╗  ██║██╔════╝\033[0;0m     ████╗ ████║██╔════╝████╗ ████║██╔═══██╗██╔══██╗╚██╗ ██╔╝
\033[2;31m██║     ██║██╔██╗ ██║█████╔╝ ██║██╔██╗ ██║██║  ███╗\033[0;0m    ██╔████╔██║█████╗  ██╔████╔██║██║   ██║██████╔╝ ╚████╔╝ 
\033[2;31m██║     ██║██║╚██╗██║██╔═██╗ ██║██║╚██╗██║██║   ██║\033[0;0m    ██║╚██╔╝██║██╔══╝  ██║╚██╔╝██║██║   ██║██╔══██╗  ╚██╔╝  
\033[2;31m███████╗██║██║ ╚████║██║  ██╗██║██║ ╚████║╚██████╔╝\033[0;0m    ██║ ╚═╝ ██║███████╗██║ ╚═╝ ██║╚██████╔╝██║  ██║   ██║   
\033[2;31m╚══════╝╚═╝╚═╝  ╚═══╝╚═╝  ╚═╝╚═╝╚═╝  ╚═══╝ ╚═════╝\033[0;0m     ╚═╝     ╚═╝╚══════╝╚═╝     ╚═╝ ╚═════╝ ╚═╝  ╚═╝   ╚═╝   
                                                                                                               
                                                                                             \033[2;31m By BooRuleDie\033[0;0m

"""
    clear_screen(0)
    print(banner)                                                                                                         

def print_menu():

    menu = """\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m

       \033[1;92m+\033[0;0m=======\033[1;92m======\033[0;0m=======\033[1;92m+\033[0;0m    \033[1;36m+\033[0;0m======\033[1;36m========\033[0;0m=====\033[1;36m+\033[0;0m    \033[1;31m+\033[0;0m=====\033[1;31m=========\033[0;0m=====\033[1;31m+\033[0;0m    \033[1;34m+\033[0;0m====\033[1;34m===========\033[0;0m====\033[1;34m+\033[0;0m
       |                    |    |                   |    |                   |    |                   |
       |    \033[1;92m1 - Start\033[0;0m       |    | \033[1;36m2 - Start & Save\033[0;0m  |    |     \033[1;31m3 - Quit\033[0;0m      |    |  \033[1;34m4 - Manual Page\033[0;0m  |
       |                    |    |                   |    |                   |    |                   |
       \033[1;92m+\033[0;0m=======\033[1;92m======\033[0;0m=======\033[1;92m+\033[0;0m    \033[1;36m+\033[0;0m======\033[1;36m========\033[0;0m=====\033[1;36m+\033[0;0m    \033[1;31m+\033[0;0m=====\033[1;31m=========\033[0;0m=====\033[1;31m+\033[0;0m    \033[1;34m+\033[0;0m====\033[1;34m===========\033[0;0m====+

\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m=\033[2;31m█\033[0;0m"""
    print(menu)
    
    # OPTION SELECTION  

def get_operation_number():
    question = """
  \033[2;33mOperation number > \033[0;0m """
    OPTION = input(question)

        # VALIDATION OF WHETHER OPTION NUMBER IS 1,2,3,4 OR NOT

    while OPTION not in ["1","2","3","4"]:

        clear_screen(0)
        print_banner()
        print_menu()

        print("\n  Invalid Entry, Operation number must be \033[1;32m1\033[0;0m, \033[1;36m2\033[0;0m, \033[1;31m3\033[0;0m or \033[1;34m4\033[0;0m.")
        OPTION = input(question)
    return OPTION

def print_manual_page():

    manual_page = """\033[2;31m+\033[0;0m===============================================================\033[2;31m+\033[0;0m
|                         \033[2;37;41m MANUAL PAGE \033[0;0;0m                         |
\033[2;31m+---\033[0;0m=========================================================\033[2;31m---+\033[0;0m
|                                                               |
|  \033[1;32m 1 - Start: \033[0;0m                                                 |
|                                                               |
|  Starts script and let you select number of words that will   |
|  be shown. Important thing is after you're done with script,  |
|  it's not going to save anything. To do it please read below. |
|                                                               |
\033[2;31m+---\033[0;0m=========================================================\033[2;31m---+\033[0;0m
|                                                               |
|  \033[1;36m 2 - Start & Save: \033[0;0m                                          |
|                                                               |
|  Perform almost same process with operation 1 but instead of  |
|  quiting directly it looks for a file named \033[1;35m"linking_method_\033[0;0m  |
|  \033[1;35mlog.xlsx"\033[0;0m. If there is such a file in directory, it saves    |
|  results there, Otherwise it creates the file then saves.     |
|                                                               |
\033[2;31m+---\033[0;0m=========================================================\033[2;31m---+\033[0;0m
|                                                               |
|  \033[1;31m 3 - Quit: \033[0;0m                                                  |
|                                                               |
|  Exits without performing any activity.                       |
|                                                               |
\033[2;31m+---\033[0;0m=========================================================\033[2;31m---+\033[0;0m
|                                                               |
|  \033[1;34m 4 - Manual Page:\033[0;0m                                            |
|                                                               |
|  Prints manual page                                           |
|                                                               |
\033[2;31m+\033[0;0m===============================================================\033[2;31m+\033[0;0m"""

    print(manual_page)

def openpyxl(save_operation_number,right,wrong):
    if save_operation_number == 1:
        
        wb = Workbook()
        ws = wb.active            
            
        ws.cell(2,2).font = Font(bold=True,color="000000FF")
        ws.cell(2,3).font = Font(bold=True,color="00FF0000")
        ws.cell(2,4).font = Font(bold=True,color="0000FF00")
        ws.cell(2,5).font = Font(bold=True,color="000000FF")

        for number in range(2,6):

            # STYLE FOR TITLE
            ws.cell(2,number).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(2,number).border = Border(left=Side(border_style="thick"),right=Side(border_style="thick"),top=Side(border_style="thick"),bottom=Side(border_style="thick"))

            #FOR VALUES
            ws.cell(3,number).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(3,number).font = Font(bold=True)
            ws.cell(3,number).border = Border(left=Side(border_style="thick"),right=Side(border_style="thick"),top=Side(border_style="thick"),bottom=Side(border_style="thick"))
        
        ws.cell(3,2).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit = True)

        # WRITING
        ws.cell(2,2).value = "Date"
        ws.cell(2,3).value = "Wrong"
        ws.cell(2,4).value = "Right"
        ws.cell(2,5).value = "Total"

        date_value = date.today().strftime("%d-%m-%Y")
        ws.cell(3,2).value = date_value
        ws.cell(3,3).value = wrong
        ws.cell(3,4).value = right
        ws.cell(3,5).value = right + wrong
            
        wb.save(filename="linking_method_log.xlsx")

    elif save_operation_number==2:

        wb = load_workbook("linking_method_log.xlsx")
        ws = wb.active

        next_row = ws.max_row + 1
        date_value = date.today().strftime("%d-%m-%Y")

        for number in range(2,6):
 
            ws.cell(next_row,number).alignment = Alignment(horizontal='center', vertical='center') 
            ws.cell(next_row,number).font = Font(bold=True)            
            ws.cell(next_row,number).border = Border(left=Side(border_style="thick"),right=Side(border_style="thick"),top=Side(border_style="thick"),bottom=Side(border_style="thick"))

        ws.cell(next_row,2).alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit = True)

        ws.cell(next_row,2).value = date_value
        ws.cell(next_row,3).value = wrong
        ws.cell(next_row,4).value = right
        ws.cell(next_row,5).value = right + wrong

        wb.save(filename="linking_method_log.xlsx")
        
def major(if_operation_number):
    LIST_THAT_WILL_BE_DISPLAYED = []
    LIST_THAT_USR_ENTERED = []

    Words = [["computer","monitor","apple","banana","car","time", "year", "people", "way", "day", "man", "thing", "woman", "life", "child", "world", "school", "state", "family", "student", "group", "country", "problem", "hand", "part", "place", "case", "week", "company", "system", "program", "question", "work", "government", "number", "night", "point", "home", "water", "room", "mother", "area", "money", "story", "fact", "month", "lot", "right", "study", "book", "eye", "job", "word", "business", "issue", "side", "kind", "head", "house", "service", "friend", "father", "power", "hour", "game", "line", "end", "member", "law", "car", "city", "community", "name", "president", "team", "minute", "idea", "kid", "body", "information", "back", "parent", "face", "others", "level", "office", "door", "health", "person", "art", "war", "history", "party", "result", "change", "morning", "reason", "research", "girl", "guy", "moment", "air", "teacher", "force", "education"],
    0,
    ["fill","fly","emphasize","probe","write","keep","understand","have","let","watch","do","begin","follow","say","seem","stop","go","help","create","can","talk","speak","get","turn","read","would","start","allow","make","might","add","know","show","spend","will","hear","grow","think","play","open","take","run","walk","see","move","win","come","like","offer","could","live","remember","want","believe","love","look","hold","consider","use","bring","appear","find","happen","buy","give","must","wait","tell","write","serve","work","provide","die","may","sit","send","should","stand","expect","call","lose","build","try","pay","stay","ask","meet","fall","need","include","cut","feel","continue","reach","become","set","kill","leave","learn","remain","put","change","mean","lead"],
    0,
    ["Protocol","FTP","SSH","VPN","base64","Antivirus", "Authenticator", "Blacklist", "Backup", "Closed Source", "Cloud", "Data Loss Prevention (DLP)", "Data encryption", "Data protection", "Exploit", "Firewall", "Group Authenticator", "Honeypot", "IP Address", "Identity check", "Incident response plan", "Insider threat", "Open Source", "Patch", "ReCAPTCHA", "Adware", "DDoS (denial-of-service attack)", "Keylogger", "Malware (malicious software)", "Ransomware", "Ransomcloud", "Free Ransomware Assessment", "Rootkit", "Spyware", "Trojan", "Virus", "Worm", "Hacking and Social Engineering", "Attack vector", "Advanced Persistent Threats", "Account Hijacking", "Bot", "Brute force attack", "Crypojacking", "Catfishing", "Cracker", "Hacker", "Phishing", "Password Sniffing", "Social Engineering"],
    0,
    ["John","Amy","Math","Elizabeth","Andrew","James", "Mary", "Robert", "Patricia", "John", "Jennifer", "Michael", "Linda", "William", "Elizabeth", "David", "Barbara", "Richard", "Susan", "Joseph", "Jessica", "Thomas", "Sarah", "Charles", "Karen", "Christopher", "Nancy", "Daniel", "Lisa", "Matthew", "Betty", "Anthony", "Margaret", "Mark", "Sandra", "Donald", "Ashley", "Steven", "Kimberly", "Paul", "Emily", "Andrew", "Donna", "Joshua", "Michelle", "Kenneth", "Dorothy", "Kevin", "Carol", "Brian", "Amanda", "George", "Melissa", "Edward", "Deborah", "Ronald", "Stephanie", "Timothy", "Rebecca", "Jason", "Sharon", "Jeffrey", "Laura", "Ryan", "Cynthia", "Jacob", "Kathleen", "Gary", "Amy", "Nicholas", "Shirley", "Eric", "Angela", "Jonathan", "Helen", "Stephen", "Anna", "Larry", "Brenda", "Justin", "Pamela", "Scott", "Nicole", "Brandon", "Emma", "Benjamin", "Samantha", "Samuel", "Katherine", "Gregory", "Christine", "Frank", "Debra", "Alexander", "Rachel", "Raymond", "Catherine", "Patrick", "Carolyn", "Jack", "Janet", "Dennis", "Ruth", "Jerry", "Maria"],
    0,
    ["Paris","Czech Republic","USA","Mexico","Netherlands","Hong Kong", "Singapore", "Bangkok", "London", "UK", "Macau", "Kuala Lumpur", "Malaysia", "Shenzhen", "New York City", "Antalya", "Paris", "France", "Istanbul", "Turkey", "Rome", "Dubai", "UAE", "Guangzhou", "Phuket", "Mecca", "Saudi Arabia", "Pattaya", "Taipei City", "Taiwan", "Prague", "Czech Republic", "Shanghai", "Las Vegas", "Miami", "Barcelona", "Moscow", "Beijing", "Los Angeles", "Budapest", "Hungary", "Vienna", "Austria", "Amsterdam", "Netherlands", "Sofia", "Bulgaria", "Madrid", "Spain", "Orlando", "Ho Chi Minh City", "Vietnam", "Lima", "Peru", "Berlin", "Germany", "Tokyo", "Japan", "Warsaw", "Poland", "Chennai", "India", "Cairo", "Egypt", "Nairobi", "Kenya", "Hangzhou", "Milan", "San Francisco", "USA", "Buenos Aires", "Argentina", "Venice", "Italy", "Mexico City", "Mexico", "Dublin", "Ireland", "Seoul", "South Korea", "Mugla", "Mumbai", "Denpasar", "Indonesia", "Delhi", "Toronto", "Canada", "Zhuhai", "St Petersburg", "Russia", "China", "Thailand"],
    0]
   
    # TAKING INPUTS FROM USER

    print("\n  Enter number of word's genre respectively as \033[1;32m'Object'\033[0;0m, \033[1;32m'Verb'\033[0;0m, \033[1;32m'Term'\033[0;0m, \033[1;32m'Name'\033[0;0m, \033[1;32m'Country&City'\033[0;0m")
    for x in range(0,10,2):

    # CONTROL OF WHETHER INPUT IS STRING OR BUGGER THAN MAX VALUE
        temp = None

        while temp == None:

            try:
                temp = int(input(f"\n  Number of Word \033[1;32m(max {len(Words[x])})\033[0;0m: "))
            
            except ValueError:
                print("\n  \033[1;31mInvalid input!\033[0;0m")
                temp = None
            
            if temp == None:    
                pass
            
            elif temp > len(Words[x]):         
                print("\n  \033[1;31mInvalid input!\033[0;0m")
                temp = None
                

        Words[x+1] = temp
    
    # ADJUSTING WORDS

    for x in range(0,10,2):
        
        while Words[x+1] > 0:
            LIST_THAT_WILL_BE_DISPLAYED.append(random.choice(Words[x]))

                # CHECKING IF THERE IS A DUPLICATE IN "LIST THAT WILL BE DISPLAYED"

            if LIST_THAT_WILL_BE_DISPLAYED.count(LIST_THAT_WILL_BE_DISPLAYED[-1]) > 1:
                
                LIST_THAT_WILL_BE_DISPLAYED.pop()
            else:
                Words[x+1] -= 1


        # SHUFFLING LIST SO SAME TYPE WORDS WON'T BE SHOWN RESPECTIVELY

    random.shuffle(LIST_THAT_WILL_BE_DISPLAYED)

    filler = input("\n  \033[2;33mPress anything when you're ready to see words > \033[0;0m")
    clear_screen(0)
    print_banner()

    for counter,word in enumerate(LIST_THAT_WILL_BE_DISPLAYED):
        print(f"\n  {counter+1}th word: \033[1;36m{word}\033[0;0m")
        filler = input()

    filler = input("\n  \033[2;33mPress enter when you are ready to repeat > \033[0;0m")
    clear_screen(0)
    print_banner()
        # USER GUESSING ALL WORDS

    for counter,x in enumerate(LIST_THAT_WILL_BE_DISPLAYED):
        user_input = input(f"\n  \033[1;37m{counter+1}th :\033[0;0m ")
        LIST_THAT_USR_ENTERED.append(user_input)

    clear_screen(0)
    print_banner()
    right = 0
    wrong = 0

    for index, ele in enumerate(LIST_THAT_USR_ENTERED):
        if LIST_THAT_WILL_BE_DISPLAYED[index].lower() != LIST_THAT_USR_ENTERED[index].lower():
            wrong += 1
            print(f"""\n  {index+1}th word: \033[1;32m{LIST_THAT_WILL_BE_DISPLAYED[index]}\033[0;0m  \t\t\t  Your answer: \033[1;31m{LIST_THAT_USR_ENTERED[index]}\033[0;0m""")
        else:
            right += 1

    print(f"=============================================================================================================\n\n  You've answered \033[1;32m{right}\033[0;0m words correct and \033[1;31m{wrong}\033[0;0m wrong. \n  Total Words: \033[1;34m{wrong+right}\033[0;0m")

    if if_operation_number == 2:
        
        filler = input("\n  \033[2;33mPress anything to save & exit > \033[0;0m")
        
        # FIRST DETECT WHETHER THERE IS A FILE NAMED 'LINKING_METHOD_LOG.xlsx' OR NOT
    
        if Path("linking_method_log.xlsx").exists():

            print("\n  [\033[1;32m✔\033[0;0m] FILE FOUND NAMED: \033[1;32m'linking_method_log.xlsx'\033[0;0m in current directory")
                
            # OPENPYXL CODES HERE --->
            openpyxl(2,right,wrong)

            time.sleep(1)
            print("\n  [\033[1;32m✔\033[0;0m] APPENDING PROCESS COMPLETED")
            time.sleep(1)
            print("\n  \033[1;31mQuiting...\033[0;0m")   
            time.sleep(1.5)
            clear_screen(0)

        else:
            print("\n  [\033[1;31m✖\033[0;0m] COULDN'T FOUND FILE NAMED: \033[1;32m'linking_method_log.xlsx'\033[0;0m in current directory")
            time.sleep(1)
            print("\n  [\033[1;32m✔\033[0;0m] NEW FILE CREATED NAMED: \033[1;32m'linking_method_log.xlsx'\033[0;0m in current directory")
            
            # OPENPYXL CODES HERE --->
            openpyxl(1,right,wrong)
                        
            time.sleep(1)
            print("\n  [\033[1;32m✔\033[0;0m] WRITING PROCESS COMPLETED")
            time.sleep(1)
            print("\n  \033[1;31mQuiting...\033[0;0m")   
            time.sleep(1.5)
            clear_screen(0)
    
    else:
        filler = input("\n  \033[2;33mPress anything to exit > \033[0;0m")
    
        print("\n  \033[1;31mQuiting without saving...\033[0;0m")
        time.sleep(1.5)
        clear_screen(0)
        
def main():

    print_banner()
    print_menu()
    
    operation_result = get_operation_number()

    if operation_result == '3':
        print("\n  \033[1;31mExiting...\033[0;0m")
        time.sleep(1.5)
        clear_screen(0)
    
    elif operation_result == '4':
        clear_screen(0)
        print_manual_page()
        filler = input("\n  \033[2;33mPress anything to exit > \033[0;0m")
        clear_screen(0)
        
        main()

    elif operation_result == '1':
        major(1)
        
    elif operation_result == '2':
        major(2)  

              
main()
